#! usr/bin/env python3

## This script is modified from PECO splitting script
## NOTE that it is essential for every excel file to have values in sample_metadata


# %% load-in
# python package for handling excel
import openpyxl
import os
import csv
# this package has a 'here' command literally like r
from pyprojroot import here




def split2csv(wb):
    for sheetName in wb.sheetnames:
        # we don't need our reference sheets
        if sheetName.startswith(('abundance summary', 'metadata', 'codes')):
            pass
        else:
            sheet = wb[sheetName]
            logging.debug('eDNA ' + sheetName)
            # open csv, filename will be region_sheet.csv
            with open(here('./rawdata/2018calvert/seine/fabs_extract/' + version + '_' + sheetName + '.csv'), 'w') as out:
                writer = csv.writer(out)
                # write each row with iter_rows
                for row in sheet.iter_rows(values_only=True):
                    # make it stop when there are no more values
                    if row[0] == None:
                        logging.debug('endsheet')
                        break
                    else:
                        writer.writerow(row)





if __name__ == '__main__':
    version = '2021Nov_FABSMaster'
    path = here('./rawdata/2018calvert/seine/raw/FABSMasterData.xlsx')
    wb = openpyxl.load_workbook(filename=path)
    split2csv(wb)

    version = '2021Nov_FABSMaster2018'
    path = here('./rawdata/2018calvert/seine/raw/FABS_masterdata_2018.xlsx')
    wb = openpyxl.load_workbook(filename=path)
    split2csv(wb)


    
